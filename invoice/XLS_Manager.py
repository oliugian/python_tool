# -*- coding: utf-8 -*-
import xlrd
import os
import shutil
from openpyxl import load_workbook
from win32com import client
all_asin_item={}
suc = 0
fal = 0

def init_all_asin():
    workbook = xlrd.open_workbook(r'.\template\ALL_ASIN.xlsx')
    # 获取所有sheet
    sheet_name = workbook.sheet_names()[0]
    # 根据sheet索引或者名称获取sheet内容
    sheet = workbook.sheet_by_name(sheet_name)
    for i in range(1, sheet.nrows):
        asin_id = sheet.cell(i, 16).value.encode('utf-8')
        item_name= sheet.cell(i, 0).value.encode('utf-8')
        all_asin_item[asin_id]=item_name



def all_invoce_go():
    # 打开文件
    workbook = xlrd.open_workbook(r'.\all_invoice.xls')
    # 获取所有sheet
    sheet_name= workbook.sheet_names()[0]
    # 根据sheet索引或者名称获取sheet内容
    sheet = workbook.sheet_by_name(sheet_name)
    global suc
    global fal
    for i in range(1,sheet.nrows):
        pay_time=sheet.cell(i,2).value.encode('utf-8')
        shop_name=sheet.cell(i,1).value.encode('utf-8')
        cus_id=sheet.cell(i,3).value.encode('utf-8')
        cus_addr=sheet.cell(i,5).value.encode('utf-8')
        trace_id=sheet.cell(i,0).value.encode('utf-8')
        product_name=sheet.cell(i,9).value.encode('utf-8')
        asin_id=sheet.cell(i,8).value.encode('utf-8')
        price=sheet.cell(i,10).value.encode('utf-8')
        price_total=sheet.cell(i,12).value.encode('utf-8')
        number=sheet.cell(i,11).value.encode('utf-8')
        all_product_cost=sheet.cell(i,14).value.encode('utf-8')
        ship_cost=sheet.cell(i,13).value.encode('utf-8')
        if all_product_cost == "":
            all_product_cost=0.00
        if ship_cost == "":
            ship_cost=0.00
        if pay_time != "":
            pay_time = pay_time.split(" ")[0]
        all_product_cost=float(all_product_cost)
        ship_cost=float(ship_cost)
        all_cost='%.2f' %float(all_product_cost+ship_cost)
        item_name=all_asin_item.get(asin_id)
        if item_name is None:
            item_name = asin_id
        if "DE" in shop_name or "de" in shop_name:
            out_put_DE_FBA(cus_id,cus_addr,item_name,trace_id,product_name,price,number,price_total,ship_cost,all_product_cost,all_cost,pay_time)
            suc = suc + 1
        elif "XTX" in shop_name:
            out_put_XTX(cus_id,cus_addr,item_name,trace_id,product_name,price,number,price_total,ship_cost,all_product_cost,all_cost,pay_time)
            suc = suc + 1
        else:
            print "shop name in "+str(i+1)+" line is " + shop_name +". please check."
            fal = fal + 1



def out_put_DE_FBA(cus_id,cus_addr,asin_id,trace_id,product_name,price,num,price_total,ship_cost,all_product_cost,all_cost,pay_time):

    new_file_name_pdf=r"./excel_result/Rechnung-DE-"+cus_id+"-"+trace_id+".pdf"

    new_file_name=r"./excel_result/Rechnung-DE-"+cus_id+"-"+trace_id+".xlsx"
    shutil.copyfile(r"./template/DE_FBA.xlsx", new_file_name)
    wb = load_workbook(new_file_name,guess_types=True)
    sheet = wb.active
    sheet["B3"] = "Datum: "+pay_time
    sheet["E5"] = cus_id
    sheet["F5"] = cus_addr
    sheet["B9"] = asin_id
    sheet["C9"] = trace_id
    #sheet["D9"] = product_name
    sheet["E9"] = price
    sheet["F9"] = num
   # sheet["G9"] = r"€"+str(price_total)
    sheet["G10"] = ship_cost
   # sheet["G11"] = r"€"+str(all_product_cost)
    #sheet["G14"] = r"€"+str(all_cost)
    wb.save(new_file_name)
    wb.close()
    #convert_to_pdf(new_file_name,new_file_name_pdf)

def out_put_XTX(cus_id,cus_addr,asin_id,trace_id,product_name,price,num,price_total,ship_cost,all_product_cost,all_cost,pay_time):
    new_file_name=r"./excel_result/Rechnung-XTX-"+cus_id+"-"+trace_id+".xlsx"
    shutil.copyfile(r"./template/XTX-Rechnung.xlsx", new_file_name)
    wb = load_workbook(new_file_name,guess_types=True)
    sheet = wb.active
    sheet["B3"] = "Datum: " + pay_time
    sheet["B11"] = cus_id
    sheet["B13"] = cus_addr
    sheet["A16"] = asin_id
    sheet["C16"] = trace_id
    #sheet["D16"] = product_name
    sheet["E16"] = price
    sheet["F16"] = num
    #sheet["G16"] = r"€"+str(price_total)
    sheet["G18"] = ship_cost
   # sheet["G20"] = r"€"+str(all_product_cost)
    wb.save(new_file_name)
    wb.close()

def convert_to_pdf(doc,pdf):
    excel = client.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(doc)
    wb.ActiveSheet.ExportAsFixedFormat(0, pdf + '.pdf')
    wb.Close()
    excel.Quit()
if __name__ == '__main__':
    try:
        init_all_asin()
        all_invoce_go()
    except Exception,e:
        print str(e)
    print "all success: "+str(suc)+" all failed: "+str(fal)
    os.system("pause")